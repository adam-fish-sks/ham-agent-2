"""
Export employees with assigned laptops but no address to Excel file
Fetches real employee data from Workwize API
"""

import os
import psycopg2
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import requests
import urllib3
from concurrent.futures import ThreadPoolExecutor, as_completed
import time

# Suppress InsecureRequestWarning
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

load_dotenv()
DATABASE_URL = os.getenv('DATABASE_URL')
WORKWIZE_KEY = os.getenv('WORKWIZE_KEY')
WORKWIZE_BASE_URL = 'https://prod-back.goworkwize.com/api/public'

def fetch_employee_from_api(employee_id):
    """Fetch real employee data from Workwize API"""
    url = f'{WORKWIZE_BASE_URL}/employees/{employee_id}'
    headers = {
        'Authorization': f'Bearer {WORKWIZE_KEY}',
        'Content-Type': 'application/json'
    }
    
    try:
        response = requests.get(url, headers=headers, verify=False, timeout=10)
        time.sleep(0.01)  # Rate limiting
        
        if response.status_code == 200:
            emp = response.json()
            # API returns data directly, not wrapped
            if emp and isinstance(emp, dict):
                return {
                    'id': str(emp.get('employee_id', employee_id)),
                    'first_name': emp.get('given_name'),
                    'last_name': emp.get('family_name'),
                    'email': emp.get('email'),
                    'status': 'inactive' if emp.get('is_deactivated') else 'active'
                }
    except Exception as e:
        print(f"Error fetching employee {employee_id}: {e}")
    
    return None

# Connect to database
conn = psycopg2.connect(DATABASE_URL)
cur = conn.cursor()

# Query for employee IDs with laptops but no address
query = """
    SELECT DISTINCT
        e.id as employee_id,
        e."firstName" as first_name,
        e."lastName" as last_name,
        e.email,
        e.status,
        COUNT(a.id) as laptop_count
    FROM employees e
    INNER JOIN assets a ON a."assignedToId" = e.id
    WHERE e."addressId" IS NULL
    AND a.status = 'deployed'
    GROUP BY e.id, e."firstName", e."lastName", e.email, e.status
    ORDER BY e.id
"""

cur.execute(query)
db_rows = cur.fetchall()
conn.close()

print(f"Found {len(db_rows)} employees with laptops but no address")
print("Fetching real employee data from Workwize API...")

# Fetch real employee data from API, fall back to database data
rows = []
with ThreadPoolExecutor(max_workers=10) as executor:
    future_to_row = {executor.submit(fetch_employee_from_api, row[0]): row for row in db_rows}
    
    completed = 0
    for future in as_completed(future_to_row):
        db_row = future_to_row[future]
        api_emp = future.result()
        
        if api_emp:
            # Use real API data
            rows.append([
                api_emp['id'],
                api_emp['first_name'],
                api_emp['last_name'],
                api_emp['email'],
                api_emp['status'],
                db_row[5]  # laptop_count from DB
            ])
        else:
            # Fall back to database data (scrubbed)
            rows.append(list(db_row))
        
        completed += 1
        if completed % 10 == 0:
            print(f"  Processed {completed}/{len(db_rows)} employees")

print(f"Found {len(rows)} employees with laptops but no address")

# Create Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Employees No Address"

# Add headers
headers = ['Employee ID', 'First Name', 'Last Name', 'Email', 'Status', 'Laptop Count']
ws.append(headers)

# Style headers
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# Add data rows
for row in rows:
    ws.append(row)

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

# Add summary
ws.append([])
ws.append([f"Total Employees: {len(rows)}"])
ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])

# Save file
output_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 
                           'employees_no_address.xlsx')
wb.save(output_path)

print(f"✓ Excel file saved to: {output_path}")
print(f"  Total employees: {len(rows)}")
print(f"  Total laptops: {sum(row[5] for row in rows)}")
